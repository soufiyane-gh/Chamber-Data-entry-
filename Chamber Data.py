import tkinter
from tkinter import ttk
from tkinter import messagebox
import os
import openpyxl




def enter_data():
    accepted = accept_var.get()
    
    if accepted=="Accepted":
        # User info
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()
        
        if firstname and lastname:
            title = title_combobox.get()
            duration = duration_spinbox.get()
            chamber = chamber_combobox.get()
            
            # Samples info
            registration_status = reg_status_var.get()
            workorder = workorder_spinbox.get()
            numofsamples = numofsamples_spinbox.get()
            
            print("First name: ", firstname, "Last name: ", lastname)
            print("Title: ", title,"Chamber: ", chamber, "Days Duration: ", duration )
            print("# Work order number: ", workorder, "# Number of samples: ", numofsamples)
            print("Registration status", registration_status)
            print("------------------------------------------")
            
            filepath = "D:\python\data.xlsx"
            
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["First Name", "Last Name", "Title", "Days Duration", "Chamber",
                           "# Work order number", "# Number of samples", "Registration status"]
                sheet.append(heading)
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([firstname, lastname, title, duration, chamber, workorder,
                          numofsamples, registration_status])
            workbook.save(filepath)
            tkinter.messagebox.showwarning(title="successfully", message="The row got add successfully..")    
        else:
            tkinter.messagebox.showwarning(title="Error", message="First name and last name are required.")
    else:
        tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms")
        registration_status = reg_status_var.get()
def delet_row():
   
    
    
        # User info
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()
        
        
        title = title_combobox.get()
        duration = duration_spinbox.get()
        chamber = chamber_combobox.get()
        row = row_delete_entry.get()
            
            # Samples info
        
        workorder = workorder_spinbox.get()
        numofsamples = numofsamples_spinbox.get()
            
           
            
        filepath = "D:\python\data.xlsx"
            
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.delete_rows(int(row))
        workbook.save(filepath)

        tkinter.messagebox.showwarning(title="successfully", message="The row got deleted successfully..")
                
       
        
window = tkinter.Tk()
window.title("Chambers Data Entry Form")




frame = tkinter.Frame(window)
frame.configure(background ='light green')
frame.pack()

# Saving User Info
user_info_frame =tkinter.LabelFrame(frame, text="User Information")
user_info_frame.grid(row= 0, column=0, padx=20, pady=10)
user_info_frame.configure(background ='light blue')

first_name_label = tkinter.Label(user_info_frame, text="First Name")
first_name_label.grid(row=0, column=0)
last_name_label = tkinter.Label(user_info_frame, text="Last Name")
last_name_label.grid(row=0, column=1)

first_name_entry = tkinter.Entry(user_info_frame)
last_name_entry = tkinter.Entry(user_info_frame)
first_name_entry.grid(row=1, column=0)
last_name_entry.grid(row=1, column=1)

row_delete_label = tkinter.Label(frame, text="Chose a row")
row_delete_label.grid(row=4, column=0)
row_delete_entry = tkinter.Entry(frame)
row_delete_entry.configure(background ='light green')
row_delete_entry.grid(row=5, column=0)

title_label = tkinter.Label(user_info_frame, text="Title")
title_combobox = ttk.Combobox(user_info_frame, values=["", "Technician.", "Engineer.", "Operator."])
title_label.grid(row=0, column=2)
title_combobox.grid(row=1, column=2)

duration_label = tkinter.Label(user_info_frame, text="Days Duration")
duration_spinbox = tkinter.Spinbox(user_info_frame, from_=0, to="infinity")
duration_label.grid(row=2, column=0)
duration_spinbox.grid(row=3, column=0)

chamber_label = tkinter.Label(user_info_frame, text="Chamber")
chamber_combobox = ttk.Combobox(user_info_frame, values=["","HF", "DH", "TC"])
chamber_label.grid(row=2, column=1)
chamber_combobox.grid(row=3, column=1)




for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)
    

# Saving samples Info
samples_frame = tkinter.LabelFrame(frame)
samples_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)
samples_frame.configure(background ='light blue')

registered_label = tkinter.Label(samples_frame, text="Registration Status")

reg_status_var = tkinter.StringVar(value="Not Registered")
registered_check = tkinter.Checkbutton(samples_frame, text="Currently Registered",
                                       variable=reg_status_var, onvalue="Registered", offvalue="Not registered")

registered_label.grid(row=0, column=0)
registered_check.grid(row=1, column=0)
registered_check.configure(background ='orange')


workorder_label = tkinter.Label(samples_frame, text= "# Work order number")
workorder_spinbox = tkinter.Spinbox(samples_frame, from_=0, to='infinity')
workorder_label.grid(row=0, column=1)
workorder_spinbox.grid(row=1, column=1)

numofsamples_label = tkinter.Label(samples_frame, text="# Number of samples")
numofsamples_spinbox = tkinter.Spinbox(samples_frame, from_=0, to="infinity")
numofsamples_label.grid(row=0, column=2)
numofsamples_spinbox.grid(row=1, column=2)

for widget in samples_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Accept terms
terms_frame = tkinter.LabelFrame(frame, text="Terms & Conditions")
terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)
terms_frame.configure(background ='light blue')

accept_var = tkinter.StringVar(value="Not Accepted")
terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
                                  variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=0, column=0)
terms_check.configure(background ='orange')

# Button
button = tkinter.Button(frame, text="Enter data", command= enter_data)
button.configure(background ='green')
button.grid(row=3, column=0, sticky="news", padx=20, pady=10)
button = tkinter.Button(frame, text="delete data", command= delet_row)
button.configure(background ='red')
button.grid(row=6, column=0, sticky="news", padx=20, pady=10)
 
window.mainloop()
